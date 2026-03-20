"""RAG Inbox — file-based knowledge folders for contextual assistance.

Users upload files into categorized folders (work/, personal/, general/).
Files are auto-extracted to text, optionally summarized by LLM, and indexed
via ToolsDNS embeddings for semantic search.

When the agent needs context (e.g., drafting a work email), it searches
the appropriate inbox folder for relevant information.

Supported formats: PDF, DOCX, XLSX, CSV, TXT, MD, JSON
"""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger

from nanobot.agent.tools.base import Tool


# ── Text extraction per file type ──

def _extract_pdf(path: Path) -> str:
    """Extract text from PDF using PyMuPDF."""
    try:
        import fitz  # pymupdf
        doc = fitz.open(str(path))
        pages = []
        for page in doc:
            pages.append(page.get_text())
        doc.close()
        return "\n\n".join(pages)
    except ImportError:
        return f"[PDF extraction requires pymupdf: pip install pymupdf] {path.name}"
    except Exception as e:
        return f"[PDF extraction failed: {e}] {path.name}"


def _extract_docx(path: Path) -> str:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
        doc = Document(str(path))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return f"[DOCX extraction requires python-docx: pip install python-docx] {path.name}"
    except Exception as e:
        return f"[DOCX extraction failed: {e}] {path.name}"


def _extract_xlsx(path: Path) -> str:
    """Extract text from XLSX using openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"## Sheet: {sheet}")
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        wb.close()
        return "\n".join(lines)
    except ImportError:
        return f"[XLSX extraction requires openpyxl: pip install openpyxl] {path.name}"
    except Exception as e:
        return f"[XLSX extraction failed: {e}] {path.name}"


def _extract_csv(path: Path) -> str:
    """Extract text from CSV."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            lines = []
            for i, row in enumerate(reader):
                if i > 200:
                    lines.append(f"... ({i}+ rows)")
                    break
                lines.append(" | ".join(row))
            return "\n".join(lines)
    except Exception as e:
        return f"[CSV extraction failed: {e}] {path.name}"


def _extract_text(path: Path) -> str:
    """Extract text from plain text files."""
    try:
        return path.read_text(encoding="utf-8", errors="replace")[:50000]
    except Exception as e:
        return f"[Text extraction failed: {e}] {path.name}"


_EXTRACTORS = {
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
    ".xls": _extract_xlsx,
    ".csv": _extract_csv,
    ".txt": _extract_text,
    ".md": _extract_text,
    ".json": _extract_text,
    ".py": _extract_text,
    ".js": _extract_text,
    ".ts": _extract_text,
    ".html": _extract_text,
}

VALID_FOLDERS = frozenset({"work", "personal", "general"})


def extract_file_text(path: Path) -> str:
    """Extract text content from a file based on its extension."""
    ext = path.suffix.lower()
    extractor = _EXTRACTORS.get(ext)
    if extractor:
        return extractor(path)
    return f"[Unsupported file type: {ext}] {path.name}"


class InboxTool(Tool):
    """Search and manage the RAG inbox folders."""

    def __init__(self, workspace: Path):
        self._workspace = workspace
        self._inbox_dir = workspace / "inbox"

    @property
    def name(self) -> str:
        return "inbox"

    @property
    def description(self) -> str:
        return (
            "Search your inbox folders (work, personal, general) for relevant information. "
            "Use before drafting emails, answering work questions, or looking up personal docs. "
            "Actions: search (find relevant content), list (show all files), info (file details)."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["search", "list", "info"],
                    "description": "Action to perform",
                },
                "query": {
                    "type": "string",
                    "description": "Search query (for search action)",
                },
                "folder": {
                    "type": "string",
                    "enum": ["work", "personal", "general", "all"],
                    "description": "Folder to search/list. Default: all",
                },
                "file_name": {
                    "type": "string",
                    "description": "File name for info action",
                },
            },
            "required": ["action"],
        }

    async def execute(
        self,
        action: str,
        query: str = "",
        folder: str = "all",
        file_name: str = "",
        **kwargs: Any,
    ) -> str:
        self._inbox_dir.mkdir(parents=True, exist_ok=True)
        for f in VALID_FOLDERS:
            (self._inbox_dir / f).mkdir(exist_ok=True)

        if action == "search":
            return await self._search(query, folder)
        elif action == "list":
            return self._list(folder)
        elif action == "info":
            return self._info(file_name, folder)
        return f"Unknown action: {action}"

    async def _search(self, query: str, folder: str) -> str:
        """Search inbox files using extracted text + ToolsDNS if available."""
        if not query:
            return "Error: query required for search"

        # Try ToolsDNS semantic search first
        try:
            from nanobot.config.loader import load_config
            config = load_config()
            td = getattr(getattr(config, "tools", None), "toolsdns", None)
            if td and td.url:
                import httpx
                prefix = f"inbox__{folder}__" if folder != "all" else "inbox__"
                resp = httpx.post(
                    f"{td.url}/v1/search",
                    json={"query": query, "top_k": 5, "threshold": 0.1, "id_prefix": prefix},
                    headers={"Authorization": f"Bearer {td.api_key}"},
                    timeout=10,
                )
                if resp.status_code == 200:
                    results = resp.json().get("results", [])
                    if results:
                        lines = [f"Found {len(results)} relevant sections:\n"]
                        for r in results:
                            score = round(r.get("score", 0) * 100, 1)
                            title = r.get("title", "")
                            content = r.get("description", r.get("content", ""))[:400]
                            source = r.get("source_info", {}).get("file_path", "")
                            lines.append(f"**{title}** ({score}% match)")
                            if source:
                                lines.append(f"Source: {source}")
                            lines.append(f"{content}\n")
                        return "\n".join(lines)
        except Exception:
            pass

        # Fallback: local text search
        return self._local_search(query, folder)

    def _local_search(self, query: str, folder: str) -> str:
        """Fallback: grep-style search through extracted text."""
        query_lower = query.lower()
        query_words = set(query_lower.split())
        matches = []

        folders = VALID_FOLDERS if folder == "all" else {folder}
        for f in folders:
            folder_path = self._inbox_dir / f
            if not folder_path.exists():
                continue
            for file_path in folder_path.iterdir():
                if file_path.is_file() and file_path.suffix.lower() in _EXTRACTORS:
                    text = extract_file_text(file_path)
                    text_lower = text.lower()
                    # Score by word overlap
                    score = sum(1 for w in query_words if w in text_lower)
                    if score > 0:
                        # Find the best matching paragraph
                        paragraphs = text.split("\n\n")
                        best_para = max(paragraphs, key=lambda p: sum(1 for w in query_words if w in p.lower()))
                        matches.append((score, f"{f}/{file_path.name}", best_para[:300]))

        if not matches:
            return f"No results found for '{query}' in {folder} inbox."

        matches.sort(key=lambda x: -x[0])
        lines = [f"Found {len(matches)} file(s) matching '{query}':\n"]
        for score, name, preview in matches[:5]:
            lines.append(f"**{name}** (relevance: {score})")
            lines.append(f"{preview}\n")
        return "\n".join(lines)

    def _list(self, folder: str) -> str:
        """List all files in inbox folders."""
        folders = VALID_FOLDERS if folder == "all" else {folder}
        lines = []
        total = 0

        for f in sorted(folders):
            folder_path = self._inbox_dir / f
            if not folder_path.exists():
                continue
            files = sorted(folder_path.iterdir())
            real_files = [p for p in files if p.is_file() and not p.name.startswith(".")]
            if real_files:
                lines.append(f"\n## {f.title()} ({len(real_files)} files)")
                for p in real_files:
                    size = p.stat().st_size
                    size_str = f"{size / 1024:.0f}KB" if size > 1024 else f"{size}B"
                    lines.append(f"- {p.name} ({size_str}, {p.suffix})")
                    total += 1

        if total == 0:
            return "Inbox is empty. Upload files via the web interface or drop them in ~/.nanobot/workspace/inbox/work/ (or personal/, general/)."

        return f"Inbox: {total} files\n" + "\n".join(lines)

    def _info(self, file_name: str, folder: str) -> str:
        """Get details about a specific file."""
        if not file_name:
            return "Error: file_name required"

        folders = VALID_FOLDERS if folder == "all" else {folder}
        for f in folders:
            path = self._inbox_dir / f / file_name
            if path.exists():
                text = extract_file_text(path)
                size = path.stat().st_size
                return (
                    f"**{f}/{file_name}** ({size / 1024:.1f}KB)\n\n"
                    f"Extracted text ({len(text)} chars):\n\n{text[:2000]}"
                )

        return f"File '{file_name}' not found in {folder} inbox."
